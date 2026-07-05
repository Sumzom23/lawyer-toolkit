#!/usr/bin/env python3
"""
GAVEL — Contract Keyword Search
=================================
Searches a folder of contracts for specific keywords or clauses.
Uses Claude AI to find and summarize relevant sections.

SETUP (run once):
    Mac:     /usr/local/bin/pip3 install anthropic pypdf openpyxl
    Windows: pip install anthropic pypdf openpyxl

USAGE:
    Mac:     python3 contract_search.py
    Windows: python contract_search.py

    The script will ask you for your API key, folder, and keywords.
    No editing required.

OUTPUT:
    An Excel spreadsheet showing which contracts contain which keywords
    and a summary of what each relevant section says.
"""

import os
import platform
from datetime import datetime


def check_libraries():
    """Check required libraries are installed."""
    missing = []
    for lib in ['anthropic', 'pypdf', 'openpyxl']:
        try:
            __import__(lib)
        except ImportError:
            missing.append(lib)

    if missing:
        print("\nERROR: Missing required libraries: " + ", ".join(missing))
        print("Install them by running:")
        if platform.system() == 'Windows':
            print(f"    pip install {' '.join(missing)}")
        else:
            print(f"    /usr/local/bin/pip3 install {' '.join(missing)}")
        return False
    return True


def get_input():
    """Ask the user for API key, folder, and keywords."""
    print("\n" + "-" * 60)

    # Get API key
    api_key = input("Enter your Anthropic API key (starts with sk-ant-): ").strip()

    # Get folder
    while True:
        folder = input("Enter the full path to the folder containing your contracts: ").strip()
        folder = folder.strip('"').strip("'")
        folder = os.path.normpath(os.path.expanduser(folder))
        if os.path.exists(folder):
            break
        print(f"  Folder not found: {folder}")

    # Get keywords
    print("\nEnter the keywords or clauses to search for.")
    print("Press Enter after each one. Type 'done' when finished.")
    keywords = []
    while True:
        kw = input(f"  Keyword {len(keywords)+1}: ").strip()
        if kw.lower() == 'done' or kw == '':
            if keywords:
                break
            print("  Please enter at least one keyword.")
        else:
            keywords.append(kw)

    # Get output filename
    default_name = f"contract_search_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    name_input = input(f"\nOutput filename (press Enter for '{default_name}'): ").strip()
    output_name = name_input if name_input else default_name
    if not output_name.lower().endswith('.xlsx'):
        output_name += '.xlsx'

    return api_key, folder, keywords, output_name


def extract_text(filepath):
    """Extract text from a PDF or text file."""
    ext = os.path.splitext(filepath)[1].lower()
    text = ""

    try:
        if ext == '.pdf':
            from pypdf import PdfReader
            reader = PdfReader(filepath)
            for page in reader.pages:
                text += page.extract_text() or ""

        elif ext in ['.txt', '.md']:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read()

        elif ext == '.docx':
            try:
                import zipfile
                import xml.etree.ElementTree as ET
                with zipfile.ZipFile(filepath) as z:
                    with z.open('word/document.xml') as doc:
                        tree = ET.parse(doc)
                        paragraphs = tree.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t')
                        text = ' '.join(p.text for p in paragraphs if p.text)
            except:
                text = ""

    except Exception as e:
        text = ""

    return text.strip()


def search_contract(client, filename, text, keywords):
    """Use Claude to search a contract for keywords and summarize findings."""
    if not text:
        return {kw: {"found": False, "summary": "Could not extract text from file"} for kw in keywords}

    # Cap text to control API costs
    text_excerpt = text[:6000]

    keywords_str = "\n".join(f"- {kw}" for kw in keywords)

    prompt = f"""You are a legal analyst reviewing a contract.

Contract: {filename}

Contract text (excerpt):
{text_excerpt}

Search for each of the following keywords or clauses:
{keywords_str}

For each keyword, respond in this exact format:
KEYWORD: [exact keyword]
FOUND: [YES or NO]
SUMMARY: [If YES: one sentence summarizing what the contract says about this. If NO: "Not found in this contract."]

Repeat this block for every keyword. Be concise and precise."""

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=500,
        messages=[{"role": "user", "content": prompt}]
    )

    response = message.content[0].text.strip()

    # Parse the response
    results = {}
    current_kw = None

    for line in response.split('\n'):
        line = line.strip()
        if line.startswith('KEYWORD:'):
            current_kw = line.replace('KEYWORD:', '').strip()
            results[current_kw] = {"found": False, "summary": ""}
        elif line.startswith('FOUND:') and current_kw:
            found_str = line.replace('FOUND:', '').strip().upper()
            results[current_kw]["found"] = found_str == "YES"
        elif line.startswith('SUMMARY:') and current_kw:
            results[current_kw]["summary"] = line.replace('SUMMARY:', '').strip()

    # Fill in any missing keywords
    for kw in keywords:
        if kw not in results:
            results[kw] = {"found": False, "summary": "Not analyzed"}

    return results


def create_excel(all_results, keywords, output_path):
    """Create the search results Excel spreadsheet."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Summary ----
    ws1 = wb.active
    ws1.title = "Summary"

    header_fill  = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    found_fill   = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    missing_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    alt_fill     = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='dddddd'),
        right=Side(style='thin', color='dddddd'),
        top=Side(style='thin', color='dddddd'),
        bottom=Side(style='thin', color='dddddd')
    )

    # Title
    ws1.merge_cells(f'A1:{get_column_letter(len(keywords)+1)}1')
    title_cell = ws1['A1']
    title_cell.value = f"CONTRACT KEYWORD SEARCH — {datetime.now().strftime('%B %d, %Y')}"
    title_cell.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 28

    # Keywords searched
    ws1.merge_cells(f'A2:{get_column_letter(len(keywords)+1)}2')
    kw_cell = ws1['A2']
    kw_cell.value = f"Keywords searched: {' | '.join(keywords)}"
    kw_cell.font = Font(name='Arial', size=9, color='666666')
    kw_cell.alignment = Alignment(horizontal='center')

    # Header row
    headers = ["Contract"] + keywords
    col_widths = [35] + [22] * len(keywords)

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws1.column_dimensions[get_column_letter(col)].width = width

    ws1.row_dimensions[3].height = 20

    # Data rows
    for i, result in enumerate(all_results, 1):
        row = i + 3
        # Contract name
        cell = ws1.cell(row=row, column=1, value=result['filename'])
        cell.font = Font(name='Arial', size=9)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = border
        if i % 2 == 0:
            cell.fill = alt_fill

        # Keyword results
        for col, kw in enumerate(keywords, 2):
            kw_result = result['results'].get(kw, {})
            found = kw_result.get('found', False)
            cell = ws1.cell(row=row, column=col, value="✓ Found" if found else "✗ Not found")
            cell.font = Font(name='Arial', size=9, bold=found)
            cell.fill = found_fill if found else missing_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        ws1.row_dimensions[row].height = 20

    ws1.freeze_panes = 'A4'

    # ---- Sheet 2: Detailed Findings ----
    ws2 = wb.create_sheet("Detailed Findings")

    detail_headers = ["Contract", "Keyword", "Found", "Summary"]
    detail_widths  = [35, 25, 10, 60]

    # Title
    ws2.merge_cells('A1:D1')
    t = ws2['A1']
    t.value = "DETAILED FINDINGS"
    t.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    t.fill = header_fill
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    for col, (header, width) in enumerate(zip(detail_headers, detail_widths), 1):
        cell = ws2.cell(row=2, column=col, value=header)
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws2.column_dimensions[get_column_letter(col)].width = width

    row = 3
    for result in all_results:
        for kw in keywords:
            kw_result = result['results'].get(kw, {})
            found = kw_result.get('found', False)
            summary = kw_result.get('summary', '')

            values = [result['filename'], kw, "Yes" if found else "No", summary]
            fill = found_fill if found else (alt_fill if row % 2 == 0 else PatternFill())

            for col, val in enumerate(values, 1):
                cell = ws2.cell(row=row, column=col, value=val)
                cell.font = Font(name='Arial', size=9)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border
                cell.fill = fill

            ws2.row_dimensions[row].height = 30
            row += 1

    ws2.freeze_panes = 'A3'

    wb.save(output_path)


def run():
    """Main function."""
    api_key, folder, keywords, output_name = get_input()

    import anthropic
    client = anthropic.Anthropic(api_key=api_key)

    # Get supported files
    supported = ['.pdf', '.docx', '.txt', '.md']
    files = sorted([
        f for f in os.listdir(folder)
        if os.path.splitext(f)[1].lower() in supported
        and not f.startswith('.')
    ])

    if not files:
        print(f"\nNo supported files found in: {folder}")
        return

    print(f"\nFound {len(files)} contracts.")
    print(f"Searching for: {', '.join(keywords)}")
    print(f"(This may take a minute)\n")

    all_results = []
    ok = bad = 0

    for i, fname in enumerate(files, 1):
        fpath = os.path.join(folder, fname)
        try:
            text = extract_text(fpath)
            results = search_contract(client, fname, text, keywords)
            all_results.append({'filename': fname, 'results': results})

            found_kws = [kw for kw, r in results.items() if r.get('found')]
            status = f"Found: {', '.join(found_kws)}" if found_kws else "No keywords found"
            print(f"  ✓ [{i}/{len(files)}] {fname} — {status}")
            ok += 1

        except Exception as e:
            print(f"  ✗ [{i}/{len(files)}] {fname} — Error: {e}")
            bad += 1

    output_path = os.path.join(folder, output_name)
    create_excel(all_results, keywords, output_path)

    print(f"\nDone! {ok} contracts searched, {bad} errors.")
    print(f"Results saved to: {output_path}")

    print()
    again = input("Run another search? (y/n): ").strip().lower()
    if again == "y":
        run()


# ---------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------
if __name__ == '__main__':
    print("=" * 60)
    print("GAVEL — Contract Keyword Search")
    print(f"Running on: {platform.system()} {platform.release()}")
    print("=" * 60)

    if not check_libraries():
        exit(1)

    run()
