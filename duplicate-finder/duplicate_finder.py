#!/usr/bin/env python3
"""
GAVEL — Duplicate Document Finder
===================================
Scans a folder of documents and finds exact and near-duplicate files.
Exact duplicates are found using file hashing (100% accurate, free).
Near-duplicates are found using Claude AI (catches renamed/slightly edited files).

SETUP (run once):
    Mac:     /usr/local/bin/pip3 install anthropic pypdf openpyxl
    Windows: pip install anthropic pypdf openpyxl

USAGE:
    Mac:     python3 duplicate_finder.py
    Windows: python duplicate_finder.py

    The script will ask you for your folder and options when you run it.
    No editing required.

OUTPUT:
    An Excel spreadsheet listing all duplicate groups found.
"""

import os
import hashlib
import platform
from datetime import datetime


def check_libraries():
    """Check required libraries are installed."""
    missing = []
    for lib in ['openpyxl']:
        try:
            __import__(lib)
        except ImportError:
            missing.append(lib)

    if missing:
        print("\nERROR: Missing required libraries: " + ", ".join(missing))
        print("Install them by running:")
        if platform.system() == 'Windows':
            print(f"    pip install {' '.join(missing)}")
        else:
            print(f"    /usr/local/bin/pip3 install {' '.join(missing)}")
        return False
    return True


def get_input():
    """Ask the user for folder and options."""
    print("\n" + "-" * 60)

    # Get folder
    while True:
        folder = input("Enter the full path to the folder to scan: ").strip()
        folder = folder.strip('"').strip("'")
        folder = os.path.normpath(os.path.expanduser(folder))
        if os.path.exists(folder):
            break
        print(f"  Folder not found: {folder}")

    # Exact only or AI near-duplicates too?
    print("\nDuplicate detection options:")
    print("  1 — Exact duplicates only (free, fast, no API key needed)")
    print("  2 — Exact + near-duplicates using Claude AI (catches renamed/edited files)")
    mode_input = input("Choose option (1 or 2, press Enter for 1): ").strip()
    use_ai = mode_input == "2"

    api_key = ""
    if use_ai:
        api_key = input("Enter your Anthropic API key (starts with sk-ant-): ").strip()

    # Output filename
    default_name = f"duplicate_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    name_input = input(f"Output filename (press Enter for '{default_name}'): ").strip()
    output_name = name_input if name_input else default_name
    if not output_name.lower().endswith('.xlsx'):
        output_name += '.xlsx'

    return folder, use_ai, api_key, output_name


def hash_file(filepath):
    """Generate MD5 hash of a file for exact duplicate detection."""
    hasher = hashlib.md5()
    try:
        with open(filepath, 'rb') as f:
            while chunk := f.read(8192):
                hasher.update(chunk)
        return hasher.hexdigest()
    except Exception:
        return None


def get_all_files(folder):
    """Get all files in a folder (non-recursive)."""
    supported = ['.pdf', '.docx', '.txt', '.md', '.xlsx', '.csv',
                 '.png', '.jpg', '.jpeg', '.doc', '.ppt', '.pptx']
    files = []
    for fname in os.listdir(folder):
        if fname.startswith('.'):
            continue
        fpath = os.path.join(folder, fname)
        if os.path.isfile(fpath):
            ext = os.path.splitext(fname)[1].lower()
            if ext in supported:
                files.append(fpath)
    return sorted(files)


def find_exact_duplicates(files):
    """Find exact duplicates using file hashing."""
    hash_map = {}
    for fpath in files:
        file_hash = hash_file(fpath)
        if file_hash:
            if file_hash not in hash_map:
                hash_map[file_hash] = []
            hash_map[file_hash].append(fpath)

    # Only return groups with more than one file
    duplicates = {h: paths for h, paths in hash_map.items() if len(paths) > 1}
    return duplicates


def extract_text_short(filepath):
    """Extract a short text excerpt from a file for AI comparison."""
    ext = os.path.splitext(filepath)[1].lower()
    text = ""
    try:
        if ext == '.pdf':
            from pypdf import PdfReader
            reader = PdfReader(filepath)
            if reader.pages:
                text = reader.pages[0].extract_text() or ""
        elif ext in ['.txt', '.md', '.csv']:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read(1000)
        elif ext == '.docx':
            import zipfile
            import xml.etree.ElementTree as ET
            with zipfile.ZipFile(filepath) as z:
                with z.open('word/document.xml') as doc:
                    tree = ET.parse(doc)
                    paragraphs = tree.findall(
                        './/{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t')
                    text = ' '.join(p.text for p in paragraphs if p.text)[:1000]
    except:
        pass
    return text[:800].strip()


def find_near_duplicates(client, files, exact_duplicate_paths):
    """Use Claude AI to find near-duplicates among non-exact-duplicate files."""
    # Filter out files already identified as exact duplicates
    exact_paths = set()
    for paths in exact_duplicate_paths.values():
        exact_paths.update(paths)

    remaining = [f for f in files if f not in exact_paths]

    if len(remaining) < 2:
        return []

    print(f"\n  Analyzing {len(remaining)} files for near-duplicates with Claude AI...")

    # Extract text from each file
    file_texts = {}
    for fpath in remaining:
        fname = os.path.basename(fpath)
        text = extract_text_short(fpath)
        file_texts[fname] = {'path': fpath, 'text': text}
        print(f"    Reading: {fname}")

    # Build a summary for Claude to compare
    file_summaries = ""
    for fname, data in file_texts.items():
        text_preview = data['text'][:200] if data['text'] else "[Could not extract text]"
        file_summaries += f"\nFILE: {fname}\nCONTENT PREVIEW: {text_preview}\n"

    prompt = f"""You are a legal document analyst. Review these documents and identify any that appear 
to be near-duplicates (same content but different filenames, slightly edited versions, or renamed copies).

{file_summaries}

List only groups of near-duplicates you find. For each group use this format:
NEAR-DUPLICATE GROUP:
- filename1.pdf
- filename2.pdf
REASON: [one sentence explaining why these appear to be duplicates]

If no near-duplicates are found, respond with: NO NEAR-DUPLICATES FOUND"""

    try:
        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=500,
            messages=[{"role": "user", "content": prompt}]
        )
        response = message.content[0].text.strip()

        # Parse response
        near_dups = []
        if "NO NEAR-DUPLICATES FOUND" in response:
            return []

        current_group = []
        current_reason = ""

        for line in response.split('\n'):
            line = line.strip()
            if line.startswith('NEAR-DUPLICATE GROUP:'):
                if current_group:
                    near_dups.append({'files': current_group, 'reason': current_reason})
                current_group = []
                current_reason = ""
            elif line.startswith('- ') and current_group is not None:
                fname = line[2:].strip()
                if fname in file_texts:
                    current_group.append(file_texts[fname]['path'])
            elif line.startswith('REASON:'):
                current_reason = line.replace('REASON:', '').strip()

        if current_group:
            near_dups.append({'files': current_group, 'reason': current_reason})

        return near_dups

    except Exception as e:
        print(f"  AI analysis error: {e}")
        return []


def create_excel(folder, exact_dups, near_dups, output_path):
    """Create the duplicate report Excel spreadsheet."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duplicate Report"

    header_fill  = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    exact_fill   = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    near_fill    = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    group_fill   = PatternFill(start_color="e8eaf6", end_color="e8eaf6", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='dddddd'),
        right=Side(style='thin', color='dddddd'),
        top=Side(style='thin', color='dddddd'),
        bottom=Side(style='thin', color='dddddd')
    )

    # Title
    ws.merge_cells('A1:D1')
    t = ws['A1']
    t.value = f"DUPLICATE DOCUMENT REPORT — {datetime.now().strftime('%B %d, %Y')}"
    t.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    t.fill = header_fill
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Summary
    ws.merge_cells('A2:D2')
    s = ws['A2']
    s.value = (f"Folder scanned: {folder}  |  "
               f"Exact duplicate groups: {len(exact_dups)}  |  "
               f"Near-duplicate groups: {len(near_dups)}")
    s.font = Font(name='Arial', size=9, color='666666')
    s.alignment = Alignment(horizontal='center')

    # Headers
    headers = ["Type", "Group", "Filename", "Notes"]
    widths   = [20, 10, 50, 40]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20

    row = 4

    # Exact duplicates
    for g_num, (file_hash, paths) in enumerate(exact_dups.items(), 1):
        for i, fpath in enumerate(paths):
            fname = os.path.basename(fpath)
            type_label = "EXACT DUPLICATE" if i == 0 else ""
            group_label = f"Group {g_num}" if i == 0 else ""
            note = "Keep this file" if i == 0 else "Consider deleting"

            for col, val in enumerate([type_label, group_label, fname, note], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = Font(name='Arial', size=9,
                                  bold=(col == 1 and i == 0))
                cell.fill = exact_fill
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border
            ws.row_dimensions[row].height = 18
            row += 1

        # Blank separator
        row += 1

    # Near duplicates
    for g_num, group in enumerate(near_dups, len(exact_dups) + 1):
        for i, fpath in enumerate(group['files']):
            fname = os.path.basename(fpath)
            type_label = "NEAR-DUPLICATE" if i == 0 else ""
            group_label = f"Group {g_num}" if i == 0 else ""
            note = group['reason'] if i == 0 else ""

            for col, val in enumerate([type_label, group_label, fname, note], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = Font(name='Arial', size=9,
                                  bold=(col == 1 and i == 0))
                cell.fill = near_fill
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border
            ws.row_dimensions[row].height = 18
            row += 1

        row += 1

    if not exact_dups and not near_dups:
        ws.merge_cells('A4:D4')
        cell = ws['A4']
        cell.value = "No duplicates found in this folder."
        cell.font = Font(name='Arial', size=10, color='666666')
        cell.alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'A4'
    wb.save(output_path)


def run():
    """Main function."""
    folder, use_ai, api_key, output_name = get_input()

    files = get_all_files(folder)
    if not files:
        print(f"\nNo supported files found in: {folder}")
        return

    print(f"\nFound {len(files)} files. Scanning for duplicates...\n")

    # Find exact duplicates
    print("Step 1: Checking for exact duplicates...")
    exact_dups = find_exact_duplicates(files)

    exact_count = sum(len(p) for p in exact_dups.values())
    print(f"  Found {len(exact_dups)} exact duplicate groups ({exact_count} files)")

    # Find near-duplicates with AI
    near_dups = []
    if use_ai and api_key:
        print("\nStep 2: Checking for near-duplicates with Claude AI...")
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        near_dups = find_near_duplicates(client, files, exact_dups)
        print(f"  Found {len(near_dups)} near-duplicate groups")

    # Save report
    output_path = os.path.join(folder, output_name)
    create_excel(folder, exact_dups, near_dups, output_path)

    print(f"\nDone!")
    print(f"  Exact duplicate groups: {len(exact_dups)}")
    print(f"  Near-duplicate groups:  {len(near_dups)}")
    print(f"  Report saved to: {output_path}")

    print()
    again = input("Scan another folder? (y/n): ").strip().lower()
    if again == "y":
        run()


# ---------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------
if __name__ == '__main__':
    print("=" * 60)
    print("GAVEL — Duplicate Document Finder")
    print(f"Running on: {platform.system()} {platform.release()}")
    print("=" * 60)

    if not check_libraries():
        exit(1)

    run()
